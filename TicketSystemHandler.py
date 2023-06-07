import requests
import pandas as pd
import datetime
from db_handler import initialize_database, get_file_path, set_file_path
from openpyxl import load_workbook
import os

def extract_order_information():
    # Get input from the user
    product_name = input("Enter the product name: ")
    display_value = input("Enter the display value (DD/MM/YY @ HH:MM - HH:MM): ")

    # Check if the file path is already linked to the product name
    file_path = get_file_path(product_name)

    if file_path is None:
        # File path is not linked to the product name, ask for user input
        file_path = input("Enter the file path to save the file: ")
        set_file_path(product_name, file_path)

    # Extract the date portion from the display value
    date_value = display_value.split(" @ ")[0]

    # Format the date value
    date_obj = datetime.datetime.strptime(date_value, "%d/%m/%y")
    formatted_date = date_obj.strftime("%d.%m.%Y")

    # WooCommerce API endpoint URL
    url = "https://klubud.dk/wp-json/wc/v3/orders"

    # Consumer Key and Consumer Secret
    consumer_key = "ck_464ac420d491684bfa3cbbf95a774f6721fcd32c"
    consumer_secret = "cs_9d0963e25610f0ac9a65370654c567840f991874"

    page = 1
    found_matches = []
    match_count = 0
    orders_per_page = 100

    while match_count < 300 and page * orders_per_page <= 5000:
        # Send GET request to the API
        response = requests.get(
            url,
            auth=(consumer_key, consumer_secret),
            params={
                "page": page,
                "per_page": orders_per_page,
                "orderby": "date",
                "order": "desc",
            },
        )

        # Check if the request was successful
        if response.status_code == 200:
            orders = response.json()

            # Extract specific information from the orders
            for order in orders:
                line_items = order['line_items']
                for item in line_items:
                    if item.get('parent_name') == product_name:
                        meta_data = item.get('meta_data', [])
                        for meta in meta_data:
                            if meta.get('display_key') == 'Dato' and meta.get('display_value') == display_value:
                                first_name = order['billing']['first_name']
                                last_name = order['billing']['last_name']
                                email = order['billing']['email']
                                phone = order['billing']['phone']
                                message = order['customer_note']
                                quantity = item['quantity']

                                # Store the matching order
                                match = {
                                    'Fornavn': first_name,
                                    'Efternavn': last_name,
                                    'Email': email,
                                    'Telefonnummer': phone,
                                    'Besked fra køber': message,
                                    'Antal billetter': quantity
                                }
                                found_matches.append(match)
                                match_count += 1

        else:
            print("Failed to retrieve orders. Status code:", response.status_code)
            return

        page += 1

        # Print status after every 10 new orders fetched
        if page % 10 == 0:
            print(f"Fetched {page * orders_per_page} orders. Found {match_count} matching orders so far.")

    if found_matches:
        # Create a DataFrame from the list of matching orders
        df = pd.DataFrame(found_matches)

        # Generate the file name
        current_date = datetime.date.today().strftime("%d.%m.%Y")
        file_name = f"{product_name.upper().replace(' ', '')}_{formatted_date}(OPDATERET D. {current_date}).xlsx"
        file_path_with_name = f"{file_path}/{file_name}"

        # Check if the Excel file already exists
        matching_files = [file for file in os.listdir(file_path) if product_name.replace(' ', '') in file]
        if matching_files:
            existing_file = matching_files[0]
            existing_file_path = os.path.join(file_path, existing_file)

            # Load the existing workbook
            workbook = load_workbook(existing_file_path)
            writer = pd.ExcelWriter(existing_file_path, engine='openpyxl')
            writer.book = workbook

            # Get the existing sheet and find the last row with data
            sheet_name = 'Sheet1'
            existing_sheet = workbook[sheet_name]
            last_row = existing_sheet.max_row

            # Append the DataFrame to the existing sheet starting from the last row
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=last_row + 1, startcol=0)

            # Save the workbook
            writer.close()
            print(f"Matching orders appended to '{existing_file_path}'")
        else:
            # Save the DataFrame to the specified file path
            df.to_excel(file_path_with_name, index=False)
            print(f"New file created: '{file_path_with_name}'")
    else:
        print("No matching order found.")

# Initialize the database (create if not exists)
initialize_database()

# Example usage
extract_order_information()
